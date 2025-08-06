from openpyxl import load_workbook
from airflow.providers.postgres.hooks.postgres import PostgresHook


pg_hook = PostgresHook(postgres_conn_id="my_postgres_conn")
conn = pg_hook.get_conn()
cursor = conn.cursor()


def read_xlsx_and_load_to_posrgtres(**kwargs):

    wb = load_workbook(filename="/opt/airflow/dags/data/users_data.xlsx")
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    data_rows = rows[1:]  # Пропускаем заголовки

    with pg_hook.get_conn() as conn:
        with conn.cursor() as cursor:
            # Создаем таблицу, если не существует
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS raw.users(
                    id serial primary key,
                    user_name text,
                    user_surname text,
                    city text,
                    age integer,
                    card_number text
                );
            """
            )

            # Вставка данных
            cursor.executemany(
                "INSERT INTO raw.users (user_name, user_surname, city, age, card_number) VALUES (%s, %s, %s, %s, %s);",
                data_rows,
            )

            # Фиксация изменений
            conn.commit()

    print("[Данные users вставились]")


def read_orders_and_load_to_posrgtres(**kwargs):

    wb = load_workbook(filename="/opt/airflow/dags/data/orders_data.xlsx")
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    data_rows = rows[1:]

    # print(data_rows)

    with pg_hook.get_conn() as conn:
        with conn.cursor() as cursor:

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS raw.orders(
                id serial primary key,
                product_name text,
                count_prod integer,
                price_per_init float,
                total_amount float,
                card_number text);
            """
            )

            sql_stmt = """INSERT INTO raw.orders (product_name, count_prod, price_per_init, total_amount, card_number)
                VALUES (%s, %s, %s, %s, %s);
            """

            for row in data_rows:
                id, name, count_prod, price_per_init, total_amount, card_number = row
                cursor.execute(
                    sql_stmt,
                    (name, count_prod, price_per_init, total_amount, card_number),
                )

            conn.commit()

    print("[Данные users вставились]")


def read_deliveries_and_load_to_posrgtres(**kwargs):
    wb = load_workbook(filename="/opt/airflow/dags/data/deliveries_data.xlsx")
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    data_rows = rows[1:]

    # print(data_rows)

    with pg_hook.get_conn() as conn:
        with conn.cursor() as cursor:

            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS raw.deliveries(
                order_number text,
                id serial primary key,
                products text,
                company_deliver text,
                amount_deliver float CHECK (amount_deliver > 0),
                name_courier text,
                tel_number text,
                start_date TIMESTAMP,
                end_date TIMESTAMP,
                city_deliver text,
                warehouse_sender text,
                warehouse_address text
                ); """
            )

            # Теперь выполняем вставку данных
            for row in data_rows:
                (
                    order_number,
                    id,
                    products,
                    company_deliver,
                    amount_deliver,
                    name_courier,
                    tel_number,
                    start_date,
                    end_date,
                    city_deliver,
                    warehouse_sender,
                    warehouse_address,
                ) = row

                cursor.execute(
                    "INSERT INTO raw.deliveries (order_number, products, company_deliver, amount_deliver, name_courier, tel_number, start_date, end_date, city_deliver, warehouse_sender, warehouse_address) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    (
                        order_number,
                        products,
                        company_deliver,
                        amount_deliver,
                        name_courier,
                        tel_number,
                        start_date,
                        end_date,
                        city_deliver,
                        warehouse_sender,
                        warehouse_address,
                    ),
                )

            conn.commit()

    print("[Данные delivers  вставились]")
